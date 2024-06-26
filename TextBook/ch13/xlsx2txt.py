import openpyxl
import sys

if len(sys.argv) < 2:
    sys.exit('Usage: python xlsx2text.py texts.xlsx')

src = sys.argv[1]

wb = openpyxl.load_workbook(src)
sheet = wb.active

for col in range(1, sheet.max_column + 1):
    filename = '{}_{:03}.txt'.format(src, col)
    text_file = open(filename, 'w', encoding='utf-8')
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(column=col, row=row).value
        if value != None:
            text_file.write(str(value))
            text_file.write('\n')
    text_file.close()
