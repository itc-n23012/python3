import openpyxl
import sys

if len(sys.argv) < 4:
    sys.exit("Usage: python blankRowInserter.py N M src.xlsx")

n, m, src = int(sys.argv[1]), int(sys.argv[2]), sys.argv[3]
wb, new_wb = openpyxl.load_workbook(src), openpyxl.Workbook()
sheet, new_sheet = wb.active, new_wb.active

for row in range(1, sheet.max_row + 1):
    new_row = row + m if row >= n else row
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(column=col, row=new_row).value = sheet.cell(column=col, row=row).value

new_wb.save(f"{src}.ins.xlsx")

